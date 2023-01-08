from __future__ import annotations

import contextlib
import datetime
import logging
import re
from dataclasses import dataclass
from enum import IntEnum
from typing import Generator

from openpyxl.worksheet.worksheet import Worksheet

import rtu_schedule_parser.utils.academic_calendar as academic_calendar
from rtu_schedule_parser import ExamsSchedule
from rtu_schedule_parser.constants import Degree, ExamType, Institute, ScheduleType
from rtu_schedule_parser.excel_formatter import ExcelFormatter
from rtu_schedule_parser.parser import ScheduleParser
from rtu_schedule_parser.schedule import Exam, ExamEmpty, ExamsSchedule
from rtu_schedule_parser.schedule_data import ScheduleData

__all__ = ["ExcelExamScheduleParser"]

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


class _ColumnDataType(IntEnum):
    """
    Used to determine the type of data in the column. The order of the values is important.

    The values are based on the column offset from the group name column:
    +-------+-------+--------+-------+-------+--------+
    | месяц | число | группа | время | № ауд | Ссылка |
    +-------+-------+--------+-------+-------+--------+
    |    -2 |    -1 |      0 |     1 |     2 |      3 |
    +-------+-------+--------+-------+-------+--------+
    """

    __slots__ = ()

    MONTH = -2
    DAY = -1
    GROUP = 0  # Group name column offset. Also used for the exam name column.
    START_TIME = 1
    ROOM = 2
    LINK = 3


@dataclass
class _ExamRow:
    """Represents a cell in the exams schedule table."""

    month: academic_calendar.Month
    day: int
    # time_start: datetime.time  # The start time of the exam
    row: tuple[str, ...]  # The row of the table


class ExcelExamScheduleParser(ScheduleParser):
    def __init__(
        self,
        document_path: str,
        period: academic_calendar.Period,
        institute: Institute,
        degree: Degree,
    ) -> None:
        super().__init__(document_path, ExcelFormatter(), period, institute, degree)

    def __parse_exams(
        self, group_column: int, exam_rows: list[_ExamRow], worksheet: Worksheet
    ) -> Generator[Exam | ExamEmpty, None, None]:
        """
        Parses the exams from the table.

        Args:
            group_column: The column in which the group name is located.
            exam_rows: The rows of the table.
            worksheet: The worksheet in which the table is located.

        Yields:
            The parsed exams.
        """

        # Convert to 0-based index
        group_column -= 1

        # column = list(worksheet.iter_cols(min_col=group_column, max_col=group_column))

        for i, exam_row_data in enumerate(exam_rows):
            row = exam_row_data.row

            exam_type, exam_name, exam_teachers, time_start = None, None, None, None
            rooms = row[group_column + _ColumnDataType.ROOM].value
            start_time_cell_value = row[group_column + _ColumnDataType.START_TIME].value

            try:
                exam_type = exam_rows[i].row[group_column + _ColumnDataType.GROUP].value
                if exam_type:
                    normalized_exam_type_or_name = exam_type.replace(" ", "").lower()
                    if normalized_exam_type_or_name == "консультация":
                        exam_type = ExamType.CONSULTATION
                    elif normalized_exam_type_or_name == "экзамен":
                        exam_type = ExamType.EXAMINATION
                    else:
                        exam_type = None

                if (
                    exam_type
                ):  # If the exam type is not None, then the exam name and teachers are in the next row
                    # Example of offsets:
                    # +---------------------------+---+
                    # | Консультация              | 0 |
                    # +---------------------------+---+
                    # | Проектирование баз данных | 1 |
                    # +---------------------------+---+
                    # | Богомольная Г.В.          | 2 |
                    # +---------------------------+---+
                    exam_name = (
                        exam_rows[i + 1].row[group_column + _ColumnDataType.GROUP].value
                    )
                    exam_teachers = (
                        exam_rows[i + 2].row[group_column + _ColumnDataType.GROUP].value
                    )
                    exam_teachers = str(exam_teachers) if exam_teachers else ""

                if exam_name:
                    exam_name = exam_name.strip()
                    exam_teachers = self._formatter.get_teachers(exam_teachers)

                if rooms:
                    rooms = self._formatter.get_rooms(rooms)
                    rooms = [self._set_default_campus(room) for room in rooms]

                if start_time_cell_value:
                    try:
                        # In one cell, the time can be written in two ways:
                        if times := re.findall(
                            r"(\d{2}-\d{2})",
                            start_time_cell_value,
                        ):
                            time_start = datetime.time(
                                hour=int(times[0].split("-")[0]),
                                minute=int(times[0].split("-")[1]),
                            )

                    except ValueError:
                        time_start = None

                if exam_name is None:
                    yield ExamEmpty(
                        month=exam_row_data.month,
                        day=exam_row_data.day,
                    )

                elif exam_name and exam_type and time_start:
                    yield Exam(
                        month=exam_row_data.month,
                        day=exam_row_data.day,
                        time_start=time_start,
                        name=exam_name,
                        exam_type=exam_type,
                        teachers=exam_teachers or [],
                        rooms=rooms or [],
                    )

            except IndexError:
                break

    def __parse_exams_rows(
        self, group_cell_index: int, group_row_index: int, worksheet: Worksheet
    ) -> Generator[_ExamRow, None, None]:
        """
        Returns a list of `_ExamRow` objects. The list contains the cells in the table that contain the exam data.
        """

        # The index of the first row in the table
        initial_row_num = group_row_index + 1

        row_count = min(worksheet.max_row, 100)

        month, day = None, None

        group_cell_index -= 1  # Convert to 0-based index

        for row in worksheet.iter_rows(min_row=initial_row_num, max_row=row_count):
            month_cell_value = row[group_cell_index + _ColumnDataType.MONTH].value
            day_cell_value = row[group_cell_index + _ColumnDataType.DAY].value

            with contextlib.suppress(ValueError):
                if month_cell_value:
                    month_cell_value = month_cell_value.replace(" ", "")
                    # Drop duplicate substrings from the month name. For example, "январьянварьянварь" -> "январь"
                    month_cell_value = re.sub(
                        r"(\w+).*\1", r"\1", month_cell_value, flags=re.IGNORECASE
                    )
                    month = academic_calendar.Month.from_str(month_cell_value)

                if day_cell_value:
                    if only_digits := "".join(filter(str.isdigit, str(day_cell_value))):
                        day = int(only_digits)

                if month and day:
                    yield _ExamRow(
                        month=month,
                        day=day,
                        row=row,
                    )

    def __parse_worksheet(
        self, worksheet: Worksheet, force: bool = False
    ) -> list[ExamsSchedule] | None:
        schedule = []  # type: list[ExamsSchedule]

        group_name_row = self._find_group_row(worksheet)

        if group_name_row is None:
            return

        group_columns = self._get_group_columns(group_name_row, worksheet)

        first_group_column = group_columns[0][1]
        exams_cells = list(
            self.__parse_exams_rows(first_group_column, group_name_row, worksheet)
        )

        for group_column in group_columns:
            try:
                exams = list(
                    self.__parse_exams(group_column[1], exams_cells, worksheet)
                )
                group_name = group_column[0]

                logger.info(
                    f"Processing group '{group_name}', worksheet '{worksheet.title}'"
                )

                schedule.append(
                    ExamsSchedule(
                        group=group_name,
                        period=self._period,
                        institute=self._institute,
                        degree=self._degree,
                        document_url=None,  # TODO: implement
                        exams=exams,
                    )
                )

            except ValueError:
                if not force:
                    raise
                else:
                    logger.error(
                        f"Error parsing schedule for group {group_column[0]}"
                        f" in worksheet {worksheet.title}. Skipping."
                    )

        return schedule

    def parse(
        self, force: bool = False, generate_dataframe: bool = False
    ) -> ScheduleData:
        """
        Args:
            force: If True, then the schedule will be parsed even if exceptions occur during parsing.
            generate_dataframe: If True, then the schedule will be converted to a pandas DataFrame. It increases the
                parsing time.
        """

        self._open_worksheets()

        schedule = []

        for worksheet in self._worksheets:
            if result := self.__parse_worksheet(worksheet, force):
                schedule.extend(result)

        return ScheduleData(
            schedule, generate_dataframe, schedule_type=ScheduleType.EXAM_SESSION
        )
