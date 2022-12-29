from __future__ import annotations

import os
from abc import ABCMeta, abstractmethod
from io import BytesIO

from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from xls2xlsx import XLS2XLSX

from rtu_schedule_parser.constants import RE_GROUP_NAME, Campus, Degree, Institute
from rtu_schedule_parser.formatter import Formatter
from rtu_schedule_parser.schedule import Room
from rtu_schedule_parser.schedule_data import ScheduleData
from rtu_schedule_parser.utils import Period


class ScheduleParser(metaclass=ABCMeta):
    """Abstract class for parsing schedule data."""

    # Campuses used by default if the campus is not specified in the schedule
    _DEFAULT_CAMPUS = {
        Institute.IIT: Campus.V_78,
        Institute.ITHT: Campus.V_86,
    }

    def __init__(
        self,
        document_path: str,
        formatter: Formatter,
        period: Period,
        institute: Institute,
        degree: Degree,
    ) -> None:
        self._document_path = document_path
        self._formatter = formatter
        self._period = period
        self._degree = degree
        self._institute = institute

        self._workbook: Workbook | None = None
        self._worksheets: list[Worksheet] | None = None

    def _open_worksheets(self):
        """Opens the workbook and all worksheets."""

        if self._document_path.endswith(".xls"):
            x2x = XLS2XLSX(self._document_path)
            self._document_path = f"{os.path.splitext(self._document_path)[0]}.xlsx"
            x2x.to_xlsx(self._document_path)

        input_excel = open(self._document_path, "rb")

        self._workbook = load_workbook(
            filename=BytesIO(input_excel.read()), read_only=True, data_only=True
        )
        self._worksheets = self._workbook.worksheets

    def _get_group_columns(
        self, group_row_index: int, worksheet: Worksheet
    ) -> list[tuple[str, int]]:
        """Returns a list of tuples containing the group name and the column index for each group in the table."""
        group_columns = []

        for row in worksheet.iter_rows(group_row_index):
            for cell in row:
                if cell and cell.value:
                    cell_value = str(cell.value).replace(" ", "")
                    if group_name := RE_GROUP_NAME.search(cell_value):
                        group_columns.append((group_name.group(1), cell.column))

        return group_columns

    def _find_group_row(self, worksheet) -> int | None:
        """Find the row containing the group name."""
        for row in worksheet.iter_rows(max_row=20, max_col=30):
            for cell in row:
                if (
                    cell
                    and cell.value
                    and RE_GROUP_NAME.match(str(cell.value.replace(" ", "")))
                ):
                    return cell.row

        return None

    def _set_default_campus(self, room: Room) -> Room:
        """Sets the campus to the default value if the campus is not specified."""
        new_room = room
        if new_room.campus is None and self._institute in self._DEFAULT_CAMPUS:
            new_room = Room(
                room.name, self._DEFAULT_CAMPUS[self._institute], room.room_type
            )

        return new_room

    @abstractmethod
    def parse(self) -> ScheduleData:
        raise NotImplementedError
