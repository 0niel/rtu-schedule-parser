import re

from openpyxl import load_workbook

from rtu_schedule_parser import ScheduleData
from rtu_schedule_parser.downloader import ScheduleDownloader

search_text = "Савка"

downloader = ScheduleDownloader()
all_docs = downloader.get_documents()

downloaded = downloader.download_all(all_docs)

print(f"Downloaded {len(downloaded)} files")

group_pattern = re.compile(r"[А-Яа-я]{4}\-\d{2}\-\d{2}")

schedules = None  # type: ScheduleData | None
for doc, doc_path, is_downloaded in downloaded:
    if doc_path.endswith(".xlsx"):
        workbook = load_workbook(doc_path)
        for sheet in workbook:
            for row in sheet.rows:
                for cell in row:
                    if (
                        isinstance(cell.value, str)
                        and search_text in cell.value
                        and cell.row <= 87
                    ):
                        row_num = cell.row
                        while row_num > 1:
                            row_num -= 1
                            group_cell = sheet.cell(row=row_num, column=cell.column - 2)
                            if isinstance(
                                group_cell.value, str
                            ) and group_pattern.match(group_cell.value):
                                print(
                                    f"Группа: {group_cell.value}, ячейка {group_cell.coordinate}. Найдено в файле {doc_path}, лист {sheet.title}, ячейка {cell.coordinate}"
                                )
                                break
                        break
